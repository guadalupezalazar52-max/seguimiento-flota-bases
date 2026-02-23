# app.py — Planificación vs Realidad (Pegado/Archivo) · Robusto · Español · Excel con gráficos
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import re
from io import StringIO, BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart

# Utils centralizadas (KPIs, topN, carga de libros XLSX)
from utils_ops import (
    compute_metrics as compute_metrics_u,
    top5_hours,
    worst_base,
    load_hourly_simple_book,  # para XLSX multi-hoja
)

# ==========================
# Apariencia
# ==========================
st.set_page_config(page_title="Plan vs Real — Operación (Pegado/Archivo)", layout="wide")
TEMPLATE = "plotly_dark"
FONT = "Inter, system-ui, Segoe UI, Roboto"

def stylize(fig, title=None, y_pct=False):
    fig.update_layout(
        template=TEMPLATE, title=title,
        font=dict(family=FONT, size=13, color="#E5E7EB"),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        legend_title_text="", margin=dict(t=45, r=10, b=30, l=10),
    )
    if y_pct: fig.update_yaxes(tickformat=".0%")
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(gridcolor="rgba(148,163,184,.25)")
    return fig

# ==========================
# Persistencia
# ==========================
DATA_DIR = Path("data"); DATA_DIR.mkdir(exist_ok=True)
BASES_DIR = DATA_DIR / "bases"; BASES_DIR.mkdir(exist_ok=True)
MERGED_CSV = DATA_DIR / "merged.csv"

def save_csv(df: pd.DataFrame, path: Path):
    df.to_csv(path, index=False, encoding="utf-8")

def load_csv(path: Path) -> pd.DataFrame|None:
    return pd.read_csv(path, encoding="utf-8") if path.exists() else None

# ==========================
# Estado
# ==========================
if "bases" not in st.session_state:    # { base: df (varias fechas) }
    st.session_state["bases"] = {}
if "merged" not in st.session_state:
    st.session_state["merged"] = pd.DataFrame()
if "_preview" not in st.session_state: # última previsualización fusionada (plan+real) para base/fecha
    st.session_state["_preview"] = pd.DataFrame()

# ============================================================
# Parser pegado texto — detección flexible de columnas (español)
# ============================================================
SYN = {
    "hora": ["hora","hr","tiempo","h"],
    "svc_plan": [
        "svc proy","servicios proy","servicios proyectados","svc plan","serv plan",
        "proyectado","planificado","plan","proy","proyectados"
    ],
    "svc_real": [
        "svc reales","servicios reales","svc real","serv real","real","reales",
        "observado","observados"
    ],
    "mov_plan": [
        "mov req","mov requeridos","mov plan","moviles plan","móviles plan","moviles requeridos",
        "req moviles","requeridos","plan moviles","movil requerido","moviles requeridas",
        "dotación plan","dotacion plan","staff plan","agentes plan","operadores plan"
    ],
    "mov_real": [
        "moviles x nomina","mov x nomina","moviles nomina","movil nomina","nómina","nomina",
        "mov reales","mov real","móviles reales","móviles real",
        "dotación","dotacion","staff","agentes","operadores","plantilla","planta",
        "dotación efectiva","dotacion efectiva"
    ],
    "coef_hs": [
        "coeficiente hs","coef hs","coef hs.","coeficiente horas","coeficiente segun hs op previstas"
    ],
    "dif_mov": [
        "dif moviles","dif mov","delta moviles","delta mov.","variacion moviles","variación móviles"
    ]
}
DEFAULT_ORDER = ["hora","svc_plan","svc_real","mov_plan","mov_real","coef_hs","dif_mov"]

def _norm(s: str) -> str:
    s = str(s).strip().lower()
    rep = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}
    for a,b in rep.items(): s = s.replace(a,b)
    s = re.sub(r"\s+"," ", s)
    return s

def _find_col(cols, aliases):
    m = { _norm(c): c for c in cols }
    for a in aliases:
        if a in m: return m[a]
    for a in aliases:
        for k,v in m.items():
            if a in k: return v
    return None

def _smart_sep(text: str) -> str:
    sample = text[:1000]
    if "\t" in sample: return "\t"
    if ";" in sample:  return ";"
    return ","

def _to_num_series(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()
    s = s.replace({"": np.nan, "nan": np.nan, "None": np.nan,
                   "#¿NOMBRE?": np.nan, "#¡NOMBRE?": np.nan, "#VALUE!": np.nan}, regex=False)
    def _fix_one(x:str):
        if x is np.nan or x is None: return np.nan
        txt = str(x)
        if "," in txt and "." in txt:
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", "")
        elif "," in txt:
            txt = txt.replace(",", ".")
        try:
            return float(txt)
        except:
            return np.nan
    return s.map(_fix_one)

def _to_time_series(s: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(s.astype(str), errors="coerce").dt.time
    except:
        return pd.to_datetime([""], errors="coerce").dt.time

@st.cache_data(show_spinner=False)
def parse_pasted_flexible(text: str) -> pd.DataFrame:
    if not text or not text.strip():
        return pd.DataFrame()
    sep = _smart_sep(text)

    df = pd.read_csv(StringIO(text), sep=sep, engine="python", dtype=str)
    cols = list(df.columns)
    hora_c = _find_col(cols, SYN["hora"])

    if hora_c is None:
        df = pd.read_csv(StringIO(text), sep=sep, header=None, engine="python", dtype=str)
        guess_hora_idx = None
        for j in range(min(10, df.shape[1])):
            val = str(df.iloc[0, j])
            if re.match(r"^\d{1,2}:\d{2}$", val.strip()):
                guess_hora_idx = j; break
        if guess_hora_idx is None:
            best, best_j = -1, None
            for j in range(df.shape[1]):
                ok = pd.to_datetime(df[j].astype(str), errors="coerce").dt.time.notna().sum()
                if ok > best:
                    best = ok; best_j = j
            guess_hora_idx = best_j

        n = df.shape[1]
        names = DEFAULT_ORDER[:n]
        df.columns = names
        if guess_hora_idx is not None and guess_hora_idx < len(names):
            names[guess_hora_idx] = "hora"
            df.columns = names

        out = pd.DataFrame()
        out["Hora"]    = _to_time_series(df.get("hora", pd.Series(dtype=str)))
        out["SvcPlan"] = _to_num_series(df.get("svc_plan", pd.Series(dtype=str)))
        out["SvcReal"] = _to_num_series(df.get("svc_real", pd.Series(dtype=str)))
        out["MovPlan"] = _to_num_series(df.get("mov_plan", pd.Series(dtype=str)))
        out["MovReal"] = _to_num_series(df.get("mov_real", pd.Series(dtype=str)))
        out["CoefHS"]  = _to_num_series(df.get("coef_hs", pd.Series(dtype=str)))
        out["DifMov_archivo"] = _to_num_series(df.get("dif_mov", pd.Series(dtype=str)))
        out = out[out["Hora"].notna()]
        out["HoraStr"] = pd.to_datetime(out["Hora"].astype(str)).dt.strftime("%H:%M")
        return out.reset_index(drop=True)

    sp_c   = _find_col(cols, SYN["svc_plan"])
    sr_c   = _find_col(cols, SYN["svc_real"])
    mp_c   = _find_col(cols, SYN["mov_plan"])
    mr_c   = _find_col(cols, SYN["mov_real"])
    coef_c = _find_col(cols, SYN["coef_hs"])
    dif_c  = _find_col(cols, SYN["dif_mov"])

    out = pd.DataFrame()
    out["Hora"]    = _to_time_series(df[hora_c])
    out["SvcPlan"] = _to_num_series(df[sp_c]) if sp_c else np.nan
    out["SvcReal"] = _to_num_series(df[sr_c]) if sr_c else np.nan
    out["MovPlan"] = _to_num_series(df[mp_c]) if mp_c else np.nan
    out["MovReal"] = _to_num_series(df[mr_c]) if mr_c else np.nan
    out["CoefHS"]  = _to_num_series(df[coef_c]) if coef_c else np.nan
    out["DifMov_archivo"] = _to_num_series(df[dif_c]) if dif_c else np.nan
    out = out[out["Hora"].notna()]
    out["HoraStr"] = pd.to_datetime(out["Hora"].astype(str)).dt.strftime("%H:%M")
    return out.reset_index(drop=True)

def merge_plan_real(plan_df: pd.DataFrame, real_df: pd.DataFrame) -> pd.DataFrame:
    # normalizo claves mínimas
    left  = plan_df[["Hora","HoraStr"]].copy()
    for c in ["SvcPlan","MovPlan","CoefHS","DifMov_archivo"]:
        left[c] = plan_df[c] if c in plan_df else np.nan

    right = real_df[["HoraStr"]].copy()
    for c in ["SvcReal","MovReal"]:
        right[c] = real_df[c] if c in real_df else np.nan

    m = pd.merge(left, right, on="HoraStr", how="outer")
    m["Hora"] = m["Hora"].fillna(pd.to_datetime(m["HoraStr"]).dt.time)

    if "SvcPlan" in plan_df and "SvcReal" in plan_df:
        m["SvcReal"] = np.where(m["SvcReal"].notna(), m["SvcReal"], plan_df.set_index("HoraStr")["SvcReal"].reindex(m["HoraStr"]).values)
    if "MovPlan" in plan_df and "MovReal" in plan_df:
        m["MovReal"] = np.where(m["MovReal"].notna(), m["MovReal"], plan_df.set_index("HoraStr")["MovReal"].reindex(m["HoraStr"]).values)

    dif_calc = m["MovReal"] - m["MovPlan"]
    m["DifMov"] = np.where(m.get("DifMov_archivo", pd.Series([np.nan]*len(m))).notna(), m["DifMov_archivo"], dif_calc)

    out = pd.DataFrame({
        "Hora": m["Hora"], "HoraStr": m["HoraStr"],
        "Servicios_Planificados": m["SvcPlan"],
        "Servicios_Reales": m["SvcReal"],
        "Moviles_Planificados": m["MovPlan"],
        "Moviles_Reales": m["MovReal"],
        "Coeficiente_HS": m["CoefHS"],
        "Dif_Moviles": m["DifMov"]
    }).sort_values("HoraStr").reset_index(drop=True)
    return out

def enrich_with_time_and_metrics(df: pd.DataFrame, fecha, base) -> pd.DataFrame:
    out = df.copy()
    out["Fecha"] = pd.to_datetime(str(fecha)).date()
    out["Base"]  = str(base).strip().upper()

    out["Fecha_dt"] = pd.to_datetime(out["Fecha"])
    iso = out["Fecha_dt"].dt.isocalendar()
    out["Año"]    = out["Fecha_dt"].dt.year
    out["Mes"]    = out["Fecha_dt"].dt.month
    out["Semana"] = iso.week
    out["Dia"]    = out["Fecha_dt"].dt.day

    out["Status"] = np.select(
        [
            out["Servicios_Planificados"].notna() & out["Servicios_Reales"].isna(),
            out["Servicios_Planificados"].isna() & out["Servicios_Reales"].notna()
        ],
        ["No ejecutado","No planificado"], default="OK"
    )

    out = compute_metrics_u(out)
    out["HoraStr"] = pd.to_datetime(out["Hora"].astype(str), errors="coerce").dt.strftime("%H:%M")
    return out

# ==========================
# Export Excel enriquecido (Resumen + Datos + 4 Gráficos)
# ==========================
def _excel_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def export_excel_pretty(df: pd.DataFrame, nombre="analisis_plan_vs_real.xlsx"):
    wb = Workbook()
    ws_res = wb.active; ws_res.title = "Resumen"
    ws_g   = wb.create_sheet("Gráficos")
    ws_d   = wb.create_sheet("Datos")

    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_hdr = PatternFill("solid", fgColor="0B1221")
    fill_blk = PatternFill("solid", fgColor="111827")
    font_hdr = Font(color="E5E7EB", bold=True, size=13)
    font_txt = Font(color="E5E7EB")

    for ws in [ws_res, ws_g, ws_d]:
        for col in range(1, 32):
            try:
                ws.column_dimensions[_excel_col(col)].width = 16
            except Exception:
                pass

    # -------- Resumen ----------
    tot_plan_s = float(pd.to_numeric(df.get("Servicios_Planificados", pd.Series()), errors="coerce").sum())
    tot_real_s = float(pd.to_numeric(df.get("Servicios_Reales", pd.Series()), errors="coerce").sum())
    tot_plan_m = float(pd.to_numeric(df.get("Moviles_Planificados", pd.Series()), errors="coerce").sum())
    tot_real_m = float(pd.to_numeric(df.get("Moviles_Reales", pd.Series()), errors="coerce").sum())

    efect = 1 - (abs(tot_real_s - tot_plan_s)/tot_plan_s) if tot_plan_s>0 else np.nan
    desv_s = (tot_real_s - tot_plan_s)/tot_plan_s*100 if tot_plan_s>0 else np.nan
    desv_m = (tot_real_m - tot_plan_m)/tot_plan_m*100 if tot_plan_m>0 else np.nan

    mape = (df["APE"].mean()*100) if "APE" in df and df["APE"].notna().any() else np.nan
    mae  = (df["AE"].mean()) if "AE" in df and df["AE"].notna().any() else np.nan
    fbias= ((df["Bias"].sum()/df["Servicios_Reales"].sum()*100) if "Bias" in df and df["Servicios_Reales"].sum()!=0 else np.nan)

    rows = [
        ["Indicador", "Valor"],
        ["Efectividad de la planificación", f"{efect:.2%}" if pd.notna(efect) else "—"],
        ["% Desvío Servicios (Real - Plan)", f"{desv_s:,.1f}%" if pd.notna(desv_s) else "—"],
        ["% Desvío Móviles (Real - Plan)", f"{desv_m:,.1f}%" if pd.notna(desv_m) else "—"],
        ["MAPE (Servicios)", f"{mape:,.1f}%" if pd.notna(mape) else "—"],
        ["MAE (Servicios)", f"{mae:,.2f}" if pd.notna(mae) else "—"],
        ["Forecast Bias (Servicios)", f"{fbias:,.1f}%" if pd.notna(fbias) else "—"],
    ]
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            c = ws_res.cell(row=i, column=j, value=val)
            c.font = font_hdr if i==1 or j==1 else font_txt
            c.fill = fill_hdr if i==1 or j==1 else fill_blk
            c.alignment = Alignment(horizontal="left")
            c.border = border

    exp_txt = (
        "Efectividad = 1 − |Real − Plan| / Plan (sobre Servicios).  "
        "Desvío % = (Real − Plan) / Plan × 100.  "
        "MAPE/MAE/Bias calculados sobre Servicios.  "
        "Datos sujetos a los filtros activos en la app."
    )
    ws_res.cell(row=len(rows)+2, column=1, value=exp_txt).font = Font(color="E5E7EB")

    # -------- Datos ----------
    cols = list(df.columns)
    for j, c in enumerate(cols, start=1):
        cell = ws_d.cell(row=1, column=j, value=c)
        cell.font = font_hdr; cell.fill = fill_hdr; cell.border = border
    for i, r in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(r, start=1):
            cell = ws_d.cell(row=i, column=j, value=v)
            cell.font = font_txt; cell.fill = fill_blk; cell.border = border

    ws_d.freeze_panes = "A2"
    ws_d.auto_filter.ref = f"A1:{_excel_col(len(cols))}{len(df)+1}"

    # -------- Gráficos ----------
    def col_idx(name): return cols.index(name)+1 if name in cols else None
    x_col = col_idx("HoraStr")
    y_svc_p = col_idx("Servicios_Planificados")
    y_svc_r = col_idx("Servicios_Reales")
    y_mov_p = col_idx("Moviles_Planificados")
    y_mov_r = col_idx("Moviles_Reales")
    y_dif_s = col_idx("Dif_Servicios")
    y_dif_m = col_idx("Dif_Moviles")
    max_row = len(df)+1

    if x_col and y_svc_p and y_svc_r:
        lc = LineChart(); lc.title = "Servicios — Plan vs Real"; lc.style = 12
        data1 = Reference(ws_d, min_col=y_svc_p, min_row=1, max_row=max_row)
        data2 = Reference(ws_d, min_col=y_svc_r, min_row=1, max_row=max_row)
        cats  = Reference(ws_d, min_col=x_col,   min_row=2, max_row=max_row)
        lc.add_data(data1, titles_from_data=True); lc.add_data(data2, titles_from_data=True)
        lc.set_categories(cats)
        ws_g.add_chart(lc, "A2")

    if x_col and y_dif_s:
        bc = BarChart(); bc.title = "Desvío Servicios — Real − Plan"; bc.style = 12
        data = Reference(ws_d, min_col=y_dif_s, min_row=1, max_row=max_row)
        cats = Reference(ws_d, min_col=x_col,  min_row=2, max_row=max_row)
        bc.add_data(data, titles_from_data=True); bc.set_categories(cats)
        ws_g.add_chart(bc, "A20")

    if x_col and y_mov_p and y_mov_r:
        lc2 = LineChart(); lc2.title = "Móviles — Plan vs Real"; lc2.style = 12
        data1 = Reference(ws_d, min_col=y_mov_p, min_row=1, max_row=max_row)
        data2 = Reference(ws_d, min_col=y_mov_r, min_row=1, max_row=max_row)
        cats  = Reference(ws_d, min_col=x_col,   min_row=2, max_row=max_row)
        lc2.add_data(data1, titles_from_data=True); lc2.add_data(data2, titles_from_data=True)
        lc2.set_categories(cats)
        ws_g.add_chart(lc2, "J2")

    if x_col and y_dif_m:
        bc2 = BarChart(); bc2.title = "Desvío Móviles — Real − Plan"; bc2.style = 12
        data = Reference(ws_d, min_col=y_dif_m, min_row=1, max_row=max_row)
        cats = Reference(ws_d, min_col=x_col,  min_row=2, max_row=max_row)
        bc2.add_data(data, titles_from_data=True); bc2.set_categories(cats)
        ws_g.add_chart(bc2, "J20")

    bio = BytesIO(); wb.save(bio)
    return bio.getvalue(), nombre

# ==========================
# Sidebar — Carga clara y guiada
# ==========================
with st.sidebar:
    st.header("Carga de datos (Plan/Real)")
    st.caption("Subí **.xlsx/.csv** o pegá tablas. Detecta encabezados en español y coma/punto decimal.")

    bases_exist = sorted(st.session_state["bases"].keys())
    base_sel = st.selectbox("Seleccioná o creá una Base", options=["(nueva)"] + bases_exist, index=0)
    base_name = st.text_input("Nombre de Base", value="" if base_sel=="(nueva)" else base_sel, help="Ej.: PROY_6001 / DEMOTOS")
    fecha_in = st.date_input("Fecha del día a analizar/guardar", value=None)

    up_file = st.file_uploader("Subir archivo (opcional)", type=["xlsx","csv"],
                               help="XLSX: se lee por hojas (cada hoja=Base). CSV: se detectan columnas automáticamente.")
    st.divider()

    st.caption("**Pegado directo** (opcional si no subís archivo)")
    with st.expander("Pegar Planificación (SVC PROY / MOV REQ)", expanded=False):
        txt_plan = st.text_area("Planificación", height=140, key="paste_plan")
    with st.expander("Pegar Realidad (SVC REALES / MOV. X NÓMINA)", expanded=False):
        txt_real = st.text_area("Realidad", height=140, key="paste_real")

    @st.cache_data(show_spinner=False)
    def _parse_cached(text: str):
        return parse_pasted_flexible(text)

    @st.cache_data(show_spinner=False)
    def _merge_cached(df_p: pd.DataFrame, df_r: pd.DataFrame):
        return merge_plan_real(df_p, df_r)

    def _leer_desde_archivo(file, fecha, base):
        """
        Devuelve (df_plan_formato_parser, df_real_formato_parser, reporte)
        con columnas: Hora, HoraStr, SvcPlan, SvcReal, MovPlan, MovReal, CoefHS, DifMov_archivo.
        """
        if file is None:
            return None, None, None
        try:
            if file.name.lower().endswith(".xlsx"):
                plan_all, real_all, rep = load_hourly_simple_book(file, fecha=str(fecha))
                # Adaptar a formato esperado por merge_plan_real (de este app)
                dfp = plan_all.rename(columns={
                    "Servicios_Planificados": "SvcPlan",
                    "Moviles_Planificados": "MovPlan"
                })[["Hora","HoraStr","SvcPlan","MovPlan"]].copy()
                dfr = real_all.rename(columns={
                    "Servicios_Reales": "SvcReal",
                    "Moviles_Reales": "MovReal"
                })[["Hora","HoraStr","SvcReal","MovReal"]].copy()
                return dfp, dfr, rep
            else:
                text = file.read().decode("utf-8", errors="ignore")
                dfp = parse_pasted_flexible(text)
                dfr = parse_pasted_flexible(text)
                return dfp, dfr, [{"sheet": "CSV", "ok": True, "base": base, "filas_plan": len(dfp), "filas_real": len(dfr)}]
        except Exception as e:
            st.warning(f"No se pudo leer el archivo: {e}")
            return None, None, None

    st.markdown("### Acciones")
    colA, colB = st.columns(2)
    with colA:
        if st.button("1) **Analizar** (fusionar Plan + Real)", help="Lee archivo o pegado; valida columnas; unifica por hora; calcula KPIs."):
            if not base_name:
                st.error("Ingresá nombre de Base."); st.stop()
            if not fecha_in:
                st.error("Elegí la Fecha."); st.stop()

            try:
                df_p, df_r, rep = _leer_desde_archivo(up_file, fecha_in, base_name)

                if df_p is None and df_r is None:
                    df_p = _parse_cached(txt_plan) if txt_plan.strip() else pd.DataFrame()
                    df_r = _parse_cached(txt_real) if txt_real.strip() else pd.DataFrame()

                if (df_p is None or df_p.empty) and (df_r is None or df_r.empty):
                    st.error("No hay datos en Plan ni en Real (archivo o pegado)."); st.stop()

                if df_p is None or df_p.empty: df_p = df_r.copy()
                if df_r is None or df_r.empty: df_r = df_p.copy()

                fused = _merge_cached(df_p, df_r)

                # Sumar por hora si hay duplicados
                fused = (
                    fused.groupby("HoraStr", as_index=False)
                         .agg({
                             "Hora": "first",
                             "Servicios_Planificados": "sum",
                             "Servicios_Reales": "sum",
                             "Moviles_Planificados": "sum",
                             "Moviles_Reales": "sum",
                             "Coeficiente_HS": "first",
                             "Dif_Moviles": "sum"
                         })
                )

                prev = enrich_with_time_and_metrics(fused, fecha_in, base_name)
                st.session_state["_preview"] = prev
                st.success(f"✅ Análisis listo — filas: {len(prev)}")
                if rep:
                    st.caption("Reporte de lectura (archivo):")
                    st.json(rep)
                st.dataframe(prev.head(24))
            except Exception as e:
                st.error(f"No se pudo leer/fusionar: {e}")

    with colB:
        if st.button("2) **Guardar día en Base**", help="Guarda este día/base en /data/bases/<base>.csv"):
            if st.session_state["_preview"].empty:
                st.info("Primero ejecutá **1) Analizar**.")
            else:
                df_prev = st.session_state["bases"].get(base_name, pd.DataFrame())
                if not df_prev.empty:
                    df_prev = df_prev[~df_prev["Fecha"].eq(pd.to_datetime(fecha_in).date())]
                    df_new = pd.concat([df_prev, st.session_state["_preview"]], ignore_index=True)
                else:
                    df_new = st.session_state["_preview"].copy()
                st.session_state["bases"][base_name] = df_new
                save_csv(df_new, BASES_DIR / f"{base_name}.csv")
                st.success(f"💾 Base '{base_name}' guardada ({len(df_new)} filas totales).")

    colC, colD = st.columns(2)
    with colC:
        if st.button("3) **Consolidar todo a /data/merged.csv**", help="Concatena todas las bases guardadas y genera /data/merged.csv"):
            dfs = [df.copy() for df in st.session_state["bases"].values() if not df.empty]
            merged = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
            st.session_state["merged"] = merged
            if not merged.empty:
                save_csv(merged, MERGED_CSV)
                st.success(f"📦 MERGED guardado: {len(merged):,} filas.")
            else:
                st.info("No hay datos para consolidar.")
    with colD:
        if st.button("4) **Limpiar memoria**", help="Vacía la memoria (no borra /data)."):
            st.session_state["bases"] = {}
            st.session_state["merged"] = pd.DataFrame()
            st.session_state["_preview"] = pd.DataFrame()
            st.success("🧹 Memoria limpia.")

# ==========================
# Área principal — Tabs + Filtros
# ==========================
st.title("Análisis de Planificación vs Realidad (pegado/archivo)")
tabs = st.tabs(["Tablero", "Análisis por Base", "Análisis Horario", "Auditoría Detallada"])

dfs_live = [df.copy() for df in st.session_state["bases"].values() if not df.empty]
merged_live = pd.concat(dfs_live, ignore_index=True) if dfs_live else pd.DataFrame()
st.session_state["merged"] = merged_live

# ---- apply_filters CORREGIDO (sin HTML entities y validando columnas) ----
def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica los filtros de la UI de forma segura:
    - Verifica que el DataFrame no esté vacío.
    - Filtra solo si la columna existe.
    - Evita KeyError cuando aún no hay datos cargados.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    d = df.copy()

    if "Base" in d.columns and bases_fil:
        d = d[d["Base"].isin(bases_fil)]

    if "Fecha" in d.columns and (fecha_fil is not None):
        d = d[d["Fecha"].eq(pd.to_datetime(fecha_fil).date())]

    if "Semana" in d.columns and semana_fil and semana_fil > 0:
        d = d[d["Semana"].eq(int(semana_fil))]

    if "Año" in d.columns and anio_fil and anio_fil > 0:
        d = d[d["Año"].eq(int(anio_fil))]

    if "Mes" in d.columns and mes_fil and mes_fil > 0:
        d = d[d["Mes"].eq(int(mes_fil))]

    if "HoraStr" in d.columns and horas_fil:
        d = d[d["HoraStr"].isin(horas_fil)]

    return d.reset_index(drop=True)

# ==========================
# TAB 1 — Tablero
# ==========================
with tabs[0]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("Pegá/subí datos y/o ajustá filtros.")
    else:
        st.subheader("Indicadores generales")
        tot_plan_s = df_f["Servicios_Planificados"].sum()
        tot_real_s = df_f["Servicios_Reales"].sum()
        tot_plan_m = df_f["Moviles_Planificados"].sum()
        tot_real_m = df_f["Moviles_Reales"].sum()

        desvio_s = (tot_real_s - tot_plan_s)/tot_plan_s*100 if tot_plan_s>0 else np.nan
        desvio_m = (tot_real_m - tot_plan_m)/tot_plan_m*100 if tot_plan_m>0 else np.nan
        efect    = 1 - (abs(tot_real_s - tot_plan_s)/tot_plan_s) if tot_plan_s>0 else np.nan

        k1,k2,k3,k4 = st.columns(4)
        k1.metric("Efectividad de la planificación", f"{efect:.1%}" if pd.notna(efect) else "—")
        k2.metric("Desvío Servicios (%)", f"{desvio_s:,.1f}%" if pd.notna(desvio_s) else "—")
        k3.metric("Desvío Móviles (%)", f"{desvio_m:,.1f}%" if pd.notna(desvio_m) else "—")
        k4.metric("Sesgo (Bias) Servicios (%)",
                  f"{(df_f['Bias'].sum()/df_f['Servicios_Reales'].sum()*100):.1f}%" if df_f["Servicios_Reales"].sum()!=0 else "—")

        g1 = df_f.groupby("HoraStr", as_index=False)[["Servicios_Planificados","Servicios_Reales","Dif_Servicios"]].sum()
        fig1 = px.bar(g1, x="HoraStr", y="Dif_Servicios", color="Dif_Servicios",
                      color_continuous_scale="RdYlGn", title="Desvío de Servicios por hora (Real − Plan)")
        stylize(fig1); st.plotly_chart(fig1, width="stretch")

        g2 = df_f.groupby("HoraStr", as_index=False)[["Moviles_Planificados","Moviles_Reales","Dif_Moviles"]].sum()
        fig2 = px.bar(g2, x="HoraStr", y="Dif_Moviles", color="Dif_Moviles",
                      color_continuous_scale="RdYlGn", title="Desvío de Móviles por hora (Real − Plan)")
        stylize(fig2); st.plotly_chart(fig2, width="stretch")

        piv = df_f.pivot_table(values="Dif_Servicios", index="Fecha", columns="HoraStr", aggfunc="sum").fillna(0)
        if not piv.empty:
            fig3 = px.imshow(piv, color_continuous_scale="RdYlGn", aspect="auto",
                             title="Mapa de calor — Desvío de servicios (Real − Plan)")
            stylize(fig3); st.plotly_chart(fig3, width="stretch")

        sub, sobre = top5_hours(df_f)
        wb = worst_base(df_f)

        c1,c2,c3 = st.columns(3)
        with c1: st.subheader("Top 5 Sub‑plan (Servicios)"); st.dataframe(sub, hide_index=True)
        with c2: st.subheader("Top 5 Sobre‑plan (Servicios)"); st.dataframe(sobre, hide_index=True)
        with c3: st.subheader("Base con mayor desvío"); st.dataframe(wb, hide_index=True)

# ==========================
# TAB 2 — Análisis por Base
# ==========================
with tabs[1]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Desvío por Base (Servicios)")
        g = df_f.groupby("Base", as_index=False)[["Servicios_Planificados","Servicios_Reales"]].sum()
        g["Desvío_%"] = np.where(g["Servicios_Planificados"]>0,
                                 (g["Servicios_Reales"]-g["Servicios_Planificados"])/g["Servicios_Planificados"]*100, np.nan)
        fig = px.bar(g, x="Base", y="Desvío_%", color="Desvío_%", color_continuous_scale="RdYlGn",
                     title="Desvío % por Base (Servicios)")
        stylize(fig); st.plotly_chart(fig, width="stretch")
        st.dataframe(g, hide_index=True)

# ==========================
# TAB 3 — Análisis Horario
# ==========================
with tabs[2]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Series por hora — Plan vs Real")
        g = df_f.groupby("HoraStr", as_index=False)[["Servicios_Planificados","Servicios_Reales",
                                                     "Moviles_Planificados","Moviles_Reales"]].sum()
        fig = px.line(g, x="HoraStr", y=["Servicios_Planificados","Servicios_Reales"], title="Servicios — Plan vs Real")
        stylize(fig); st.plotly_chart(fig, width="stretch")
        figm = px.line(g, x="HoraStr", y=["Moviles_Planificados","Moviles_Reales"], title="Móviles — Plan vs Real")
        stylize(figm); st.plotly_chart(figm, width="stretch")

        st.subheader("Tabla horaria (detalle)")
        st.dataframe(
            df_f[["Fecha","HoraStr","Base","Servicios_Planificados","Servicios_Reales","Dif_Servicios",
                  "Moviles_Planificados","Moviles_Reales","Dif_Moviles","Efectividad","Clasificacion"]]
            .sort_values(["Fecha","HoraStr","Base"]),
            hide_index=True
        )

# ==========================
# TAB 4 — Auditoría Detallada (descarga Excel con gráficos)
# ==========================
with tabs[3]:
    df_f = apply_filters(merged_live)
    if df_f.empty:
        st.info("No hay datos para los filtros seleccionados.")
    else:
        st.subheader("Auditoría (lo que estás viendo)")

        cols = ["Fecha","HoraStr","Base",
                "Servicios_Planificados","Servicios_Reales","Dif_Servicios","Desvio_Servicios_%",
                "Moviles_Planificados","Moviles_Reales","Dif_Moviles","Desvio_Moviles_%",
                "Efectividad","Clasificacion","Status","Semana","Mes","Año","Coeficiente_HS","Bias","AE","APE"]
        cols = [c for c in cols if c in df_f.columns]
        df_aud = df_f[cols].sort_values(["Fecha","HoraStr","Base"])
        st.dataframe(df_aud, hide_index=True)

        g_agg = df_f.groupby("HoraStr", as_index=False)[
            ["Servicios_Planificados","Servicios_Reales","Moviles_Planificados","Moviles_Reales","Dif_Servicios","Dif_Moviles"]
        ].sum().sort_values("HoraStr")

        st.markdown("##### Servicios — Plan vs Real")
        fig_svc = px.line(g_agg, x="HoraStr", y=["Servicios_Planificados","Servicios_Reales"],
                          title="Servicios: Plan vs Real (serie horaria)")
        stylize(fig_svc); st.plotly_chart(fig_svc, width="stretch")

        st.markdown("##### Desvío por hora (Servicios)")
        fig_ds = px.bar(g_agg, x="HoraStr", y="Dif_Servicios", color="Dif_Servicios",
                        color_continuous_scale="RdYlGn",
                        title="Desvío (Servicios) por hora — Real − Plan")
        stylize(fig_ds); st.plotly_chart(fig_ds, width="stretch")

        st.markdown("##### Móviles — Plan vs Real")
        fig_mov = px.line(g_agg, x="HoraStr", y=["Moviles_Planificados","Moviles_Reales"],
                          title="Móviles: Plan vs Real (serie horaria)")
        stylize(fig_mov); st.plotly_chart(fig_mov, width="stretch")

        st.markdown("##### Desvío por hora (Móviles)")
        fig_dm = px.bar(g_agg, x="HoraStr", y="Dif_Moviles", color="Dif_Moviles",
                        color_continuous_scale="RdYlGn",
                        title="Desvío (Móviles) por hora — Real − Plan")
        stylize(fig_dm); st.plotly_chart(fig_dm, width="stretch")

        xls, fname = export_excel_pretty(df_aud, "analisis_plan_vs_real.xlsx")
        st.download_button("⬇️ Descargar Excel de Auditoría (con gráficos)", data=xls, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
